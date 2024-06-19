import openpyxl
from loguru import logger
file_path = "./excel_result";

# juho : 금액을 천 단위로 내림하는 함수
def round_down(amount):
    rounded = (amount // 1000) * 1000
    remainder = amount % 1000
    return rounded, remainder
# juho : 엑셀 파일을 생성하는 함수
def exportExcel(excel_data, comment):
    logger.info('Exporting to excel file..')
    #region juho : 엑셀 데이터 계산
    excel_data['공구손료'] = round(excel_data['직접노무비'] * 0.03)
    excel_data['산재보험료'] = round(excel_data['노무비'] * excel_data['산재보험료계수'])
    excel_data['고용보험료'] = round(excel_data['노무비'] * excel_data['고용보험료계수'])
    excel_data['기타경비'] = round((excel_data['재료비']+excel_data['노무비']) * excel_data['기타경비계수'])
    excel_data['경비'] = excel_data['공구손료'] + excel_data['산재보험료'] + excel_data['고용보험료'] + excel_data['기타경비']
    excel_data['일반관리비'] = round((excel_data['노무비']+excel_data['재료비']+excel_data['경비']) * excel_data['일반관리비계수'])
    excel_data['이윤일반공사'] = round((excel_data['노무비']+excel_data['경비']+excel_data['일반관리비']) * excel_data['이윤일반공사계수'])
    excel_data['이윤간이공사'] = round((excel_data['노무비']+excel_data['경비']+excel_data['일반관리비']) * excel_data['이윤간이공사계수'])
    final = {
        '추정금액' : 0,
        '도급분(이윤15%)' : 0,
        '도급분(이윤10%)' : 0,
        '도급부가세' : 0,
        '간이단수' : 0,
        '일반단수' : 0,
    }
    final['도급분(이윤10%)'] = excel_data['재료비'] + excel_data['노무비'] + excel_data['경비'] + excel_data['일반관리비'] + excel_data['이윤간이공사']
    rounded, remainder = round_down(final['도급분(이윤10%)'])
    final['도급분(이윤10%)'] = rounded
    final['간이단수'] = remainder
    excel_data['소계'] = final['도급분(이윤10%)']
    final['도급분(이윤15%)'] = excel_data['재료비'] + excel_data['노무비'] + excel_data['경비'] + excel_data['일반관리비'] + excel_data['이윤일반공사']
    rounded, remainder = round_down(final['도급분(이윤15%)'])
    final['도급분(이윤15%)'] = rounded
    final['일반단수'] = remainder
    final['도급부가세'] = round(final['도급분(이윤10%)'] * 0.1)
    excel_data['부가세'] = final['도급부가세']
    final['추정금액'] = final['도급분(이윤10%)'] + final['도급부가세']
    excel_data['합계'] = final['추정금액']
    #endregion
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()

    #region juho : 시트 기본 항목, 제목 등 설정
    wb.active.title = "간이공사산출명세서";
    #a2 - w2 merge
    ws = wb.active
    ws.merge_cells('A2:W2')
    # insert data to a2
    # bold, fontsize 24, text align center
    ws['A2'].font = openpyxl.styles.Font(size=24, bold=True)
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws['A2'] = '간이공사비 산출명세서'
    # insert data to a3 
    # bold, fontsize 16
    ws['A3'].font = openpyxl.styles.Font(size=16, bold=True)
    ws['A3'] = 'o 공사명 : 보호계전기 데이터 취득시스템(PDAS) 포인트 증설공사'
    #endregion
    # juho : 산출명세서가 이루어지는 기본적인 항목들은 미리 스타일과, 셀위치를 정의해둠
    defaultData = [
        {'data' : '구      분', 's' : 'a4', 'e' : 'c4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '재료비', 's' : 'd4', 'e' : 'e4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '노무비', 's' : 'f4', 'e' : 'g4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '경 비', 's' : 'h4', 'e' : 'i4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '일반관리비', 's' : 'j4', 'e' : 'k4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '이 윤', 's' : 'l4', 'e' : 'm4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '소  계', 's' : 'n4', 'e' : 'o4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '부가세', 's' : 'p4', 'e' : 'r4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '합계', 's' : 's4', 'e' : 't4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '작성자', 's' : 'u4', 'e' : 'v4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '검토자', 's' : 'w4', 'e' : 'w4', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '설\n계\n공\n사\n비', 's' : 'a5', 'e' : 'a7', 'fontsize' : 9, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '회사분', 's' : 'b5', 'e' : 'c5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'd5', 'e' : 'e5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'f5', 'e' : 'g5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'h5', 'e' : 'i5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'j5', 'e' : 'k5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'l5', 'e' : 'm5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'n5', 'e' : 'o5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 'p5', 'e' : 'r5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '-', 's' : 's5', 'e' : 't5', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 넣어야됨
        {'data' : '도급분', 's' : 'b6', 'e' : 'c6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : excel_data['재료비'], 's' : 'd6', 'e' : 'e6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'], 's' : 'f6', 'e' : 'g6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['경비'], 's' : 'h6', 'e' : 'i6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['일반관리비'], 's' : 'j6', 'e' : 'k6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['이윤'], 's' : 'l6', 'e' : 'm6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['소계'], 's' : 'n6', 'e' : 'o6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['부가세'], 's' : 'p6', 'e' : 'r6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['합계'], 's' : 's6', 'e' : 't6', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        #넣어야됨
        {'data' : '계', 's' : 'b7', 'e' : 'c7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : excel_data['재료비'], 's' : 'd7', 'e' : 'e7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'], 's' : 'f7', 'e' : 'g7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['경비'], 's' : 'h7', 'e' : 'i7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['일반관리비'], 's' : 'j7', 'e' : 'k7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['이윤'], 's' : 'l7', 'e' : 'm7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['소계'], 's' : 'n7', 'e' : 'o7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['부가세'], 's' : 'p7', 'e' : 'r7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['합계'], 's' : 's7', 'e' : 't7', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '팀 원 \n이현영\n\n(인)', 's' : 'u5', 'e' : 'v7', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '차 장 \n이현영\n\n(인)', 's' : 'w5', 'e' : 'w7', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '예정가격(결정권자)', 's' : 'a8', 'e' : 'd9', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '                                         (원)  - 부가세 포함', 's' : 'e8', 'e' : 'o9', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '부      장      이 영 성             (인)', 's' : 'p8', 'e' : 'w9', 'fontsize' : 11, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '공   사   비   예   산   서', 's' : 'a10', 'e' : 'o10', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '공 사 비 정 산 서', 's' : 'p10', 'e' : 'w10', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},    
        {'data' : 'No.', 's' : 'a11', 'e' : 'b11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '품 명', 's' : 'c11', 'e' : 'd11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '규격', 's' : 'e11', 'e' : 'g11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '단 위', 's' : 'h11', 'e' : 'h11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '수 량', 's' : 'i11', 'e' : 'j11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '단 가	', 's' : 'k11', 'e' : 'l11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '금   액', 's' : 'm11', 'e' : 'n11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '비  고', 's' : 'o11', 'e' : 'o11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '수량', 's' : 'p11', 'e' : 'q11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '단 가', 's' : 'r11', 'e' : 's11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '금  액	', 's' : 't11', 'e' : 'v11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},
        {'data' : '증감', 's' : 'w11', 'e' : 'w11', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFC0'},

        # 재료비 종류별로 여러개
        {'data' : '1. 재료비', 's' : 'a12', 'e' : 'd12', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '도급분', 's' : 'c13', 'e' : 'd13', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 1. 재료비 금액
        {'data' : excel_data['재료비'], 's' : 'm12', 'e' : 'n12', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        # 1. 재료비 도급분 금액
        {'data' : excel_data['재료비'], 's' : 'm13', 'e' : 'n13', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        # 넣어야됨 노무비 종류별로 여러개
        {'data' : '2. 노무비', 's' : 'a20', 'e' : 'd20', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['노무비'], 's' : 'm20', 'e' : 'n20', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        # 직접노무비
        {'data' : '가. 직접노무비', 's' : 'c21', 'e' : 'd21', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['직접노무비'], 's' : 'm21', 'e' : 'n21', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        #간접노무비
        {'data' : '나. 간접노무비', 's' : 'c30', 'e' : 'd30', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['간접노무비'], 's' : 'm30', 'e' : 'n30', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},

        # 직접노무비
        # 넣어야됨 간접노무비
        {'data' : '직접노무비 * 13.0%', 's' : 'c31', 'e' : 'd31', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : '0.13', 's' : 'i31', 'e' : 'j31', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['직접노무비'], 's' : 'k31', 'e' : 'l31', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['간접노무비'], 's' : 'm31', 'e' : 'n31', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 넣어야됨 경비
        {'data' : '3. 경비', 's' : 'a33', 'e' : 'd33', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['경비'], 's' : 'm33', 'e' : 'n33', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        # 도급분
        {'data' : '가. 도급분', 's' : 'c34', 'e' : 'd34', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['경비'], 's' : 'm34', 'e' : 'n34', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        # 공구손료
        {'data' : ' 공구손료', 's' : 'c35', 'e' : 'd35', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '직접노무비(노임할증,품할증,토목,건축,기계,시험 공량분 제외) x 3%', 's' : 'e35', 'e' : 'l35', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '0.03', 's' : 'i36', 'e' : 'j36', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['직접노무비'], 's' : 'k36', 'e' : 'l36', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['공구손료'], 's' : 'm36', 'e' : 'n36', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 산재보험료
        {'data' : ' 산재보험료', 's' : 'c37', 'e' : 'd37', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '노무비 * ' + str(excel_data['산재보험료계수']*100) + '%', 's' : 'e37', 'e' : 'l37', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : excel_data['산재보험료계수'], 's' : 'i38', 'e' : 'j38', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'], 's' : 'k38', 'e' : 'l38', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['산재보험료'], 's' : 'i38', 'e' : 'k38', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 고용보험료
        {'data' : ' 고용보험료', 's' : 'c39', 'e' : 'd39', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '노무비 * ' + str(excel_data['고용보험료계수'] * 100) + '%', 's' : 'e39', 'e' : 'l39', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : excel_data['고용보험료계수'], 's' : 'i40', 'e' : 'j40', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'], 's' : 'k40', 'e' : 'l40', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['고용보험료'], 's' : 'i40', 'e' : 'k40', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 기타경비
        {'data' : ' 기타경비', 's' : 'c41', 'e' : 'd41', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '(도급재료비 + 노무비) * ' +  str(excel_data['기타경비계수'] * 100) + '%', 's' : 'e41', 'e' : 'l41', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : excel_data['기타경비계수'], 's' : 'i42', 'e' : 'j42', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['재료비'] + excel_data['노무비'], 's' : 'k42', 'e' : 'l42', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['기타경비'], 's' : 'i42', 'e' : 'k42', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        # 넣어야됨 일반관리비
        {'data' : '4. 일반관리비', 's' : 'a44', 'e' : 'd44', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '(도급재료비 + 노무비 + 도급경비) * ' + str(excel_data['일반관리비계수']) + '%', 's' : 'e44', 'e' : 'l44', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['일반관리비'], 's' : 'm44', 'e' : 'n44', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        {'data' : excel_data['일반관리비계수'], 's' : 'i45', 'e' : 'j45', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['재료비'] + excel_data['노무비'] + excel_data['경비'], 's' : 'k45', 'e' : 'l45', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['일반관리비'], 's' : 'i45', 'e' : 'k45', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},

        # 넣어야됨 이윤
        {'data' : '5. 이윤', 's' : 'a46', 'e' : 'd46', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : '(노무비 + 도급경비 + 일반관리비) * 15%', 's' : 'e46', 'e' : 'l46', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        # 금액
        {'data' : excel_data['이윤일반공사'], 's' : 'm46', 'e' : 'n46', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        {'data' : '일반공사', 's' : 'o46', 'e' : 'o46', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        #
        {'data' : excel_data['이윤일반공사계수'], 's' : 'i47', 'e' : 'j47', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'] + excel_data['경비'] + excel_data['일반관리비'], 's' : 'k47', 'e' : 'l47', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['이윤일반공사'], 's' : 'i47', 'e' : 'k47', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        #
        {'data' : '(노무비 + 도급경비 + 일반관리비) * 10%', 's' : 'e49', 'e' : 'l49', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'},
        {'data' : excel_data['이윤간이공사'], 's' : 'm49', 'e' : 'n49', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFF00'},
        {'data' : '간이공사', 's' : 'o49', 'e' : 'o49', 'fontsize' : 10, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        #
        {'data' : excel_data['이윤일반공사계수'], 's' : 'i50', 'e' : 'j50', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['노무비'] + excel_data['경비'] + excel_data['일반관리비'], 's' : 'k50', 'e' : 'l50', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : excel_data['이윤간이공사'], 's' : 'i50', 'e' : 'k50', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        #
        {'data' : '단수정리', 's' : 'k51', 'e' : 'l51', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '(₩'+str(final['일반단수'])+')', 's' : 'i51', 'e' : 'k51', 'fontsize' : 12, 'fontcolor' : 'f63738', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '일반공사', 's' : 'o51', 'e' : 'o51', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        #
        {'data' : '단수정리', 's' : 'k52', 'e' : 'l52', 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '(₩'+str(final['간이단수'])+')', 's' : 'i52', 'e' : 'k52', 'fontsize' : 12, 'fontcolor' : 'f63738', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'},
        {'data' : '간이공사', 's' : 'o52', 'e' : 'o52', 'fontsize' : 12, 'fontcolor' : 'f63738', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        #결과
        {'data' : '추 정 금 액', 's' : 'e53', 'e' : 'j53', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : final['추정금액'], 's' : 'k53', 'e' : 'n53', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFF00'},
        {'data' : '도 급 분(일반공사 - 이윤 15%)', 's' : 'e54', 'e' : 'j54', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : final['도급분(이윤15%)'], 's' : 'k54', 'e' : 'n54', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFF00'},
        {'data' : '도 급 분(간이공사 - 이윤 10%)', 's' : 'e55', 'e' : 'j55', 'fontsize' : 14, 'fontcolor' : 'f63738', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : final['도급분(이윤10%)'], 's' : 'k55', 'e' : 'n55', 'fontsize' : 14, 'fontcolor' : 'f63738', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFF00'},
        {'data' : '도급부가세', 's' : 'e56', 'e' : 'j56', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'},
        {'data' : final['도급부가세'], 's' : 'k56', 'e' : 'n56', 'fontsize' : 14, 'fontcolor' : '000000', 'fontstyle' : 'bold', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFF00'},
    ]

    # juho : 재료비 항목들 반복
    재료_s = 14
    for i in range(0, len(excel_data['재료'])):
        defaultData.append({'data' : ' '+excel_data['재료'][i]['품명'], 's' : 'c' + str(재료_s + i), 'e' : 'd' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'})
        defaultData.append({'data' : '', 's' : 'e' + str(재료_s + i), 'e' : 'g' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'})
        defaultData.append({'data' : 'M', 's' : 'h' + str(재료_s + i), 'e' : 'h' + str(재료_s + i), 'fontsize' : 15, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'})
        defaultData.append({'data' : excel_data['재료'][i]['수량'], 's' : 'i' + str(재료_s + i), 'e' : 'j' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : excel_data['재료'][i]['단가'], 's' : 'k' + str(재료_s + i), 'e' : 'l' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : round(excel_data['재료'][i]['수량'] * excel_data['재료'][i]['단가']), 's' : 'm' + str(재료_s + i), 'e' : 'n' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : '견적', 's' : 'o' + str(재료_s + i), 'e' : 'o' + str(재료_s + i), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
    # juho : 노무비 항목들 반복
    노무_s = 22
    for i in range(0, len(excel_data['노무'])):
        defaultData.append({'data' : ' '+excel_data['노무'][i]['품명'], 's' : 'c' + str(i+노무_s), 'e' : 'd' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'left', 'background' : 'FFFFFF'})
        defaultData.append({'data' : '', 's' : 'e' + str(i+노무_s), 'e' : 'g' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'})
        defaultData.append({'data' : 'M', 's' : 'h' + str(i+노무_s), 'e' : 'h' + str(i+노무_s), 'fontsize' : 15, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'center', 'background' : 'FFFFFF'})
        defaultData.append({'data' : excel_data['노무'][i]['수량'], 's' : 'i' + str(i+노무_s), 'e' : 'j' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : excel_data['노무'][i]['단가'], 's' : 'k' + str(i+노무_s), 'e' : 'l' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : round(excel_data['노무'][i]['수량'] * excel_data['노무'][i]['단가']), 's' : 'm' + str(i+노무_s), 'e' : 'n' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
        defaultData.append({'data' : '품'+str(i), 's' : 'o' + str(i+노무_s), 'e' : 'o' + str(i+노무_s), 'fontsize' : 12, 'fontcolor' : '000000', 'fontstyle' : '', 'borderline' : True, 'textalign' : 'right', 'background' : 'FFFFFF'})
    
    # juho : 결합된 데이터를 반복문으로, 셀하나씩 엑셀에 적용
    for d in defaultData:
        ws.merge_cells(d['s'] + ':' + d['e'])
        ws[d['s']].font = openpyxl.styles.Font(size=d['fontsize'], bold=d['fontstyle'] == 'bold', color= d['fontcolor'])
        # font color
        ws[d['s']].alignment = openpyxl.styles.Alignment(horizontal=d['textalign'])
        ws[d['s']].fill = openpyxl.styles.PatternFill(start_color=d['background'], end_color=d['background'], fill_type="solid")
        #border s to e all cells
        if d['borderline']:
            for i in range(ord(d['s'][0]), ord(d['e'][0]) + 1):
                for j in range(int(d['s'][1:]), int(d['e'][1:]) + 1):
                    ws[chr(i) + str(j)].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

        ws[d['s']] = d['data']

    
    #comment
    c_sheet = wb.create_sheet(index=0)
    c_sheet.title = "Comment"
    ws = c_sheet
    ws.merge_cells('A2:W2')
    ws['A2'] = "comment"
    ws['A2'].font = openpyxl.styles.Font(size=24, bold=True, color='000000')
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws['A2'].fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
    ws.merge_cells('A3:W8')
    ws['A3'].font = openpyxl.styles.Font(size=16, bold=False, color='000000')
    ws['A3'] = comment

    new_filename = 'new_costs.xlsx'

    wb.save(file_path + "/" + new_filename)
    logger.info("new excel file created" + new_filename)
    return new_filename
